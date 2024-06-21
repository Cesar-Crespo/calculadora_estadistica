import openpyxl

class ModeloDatos:
    def __init__(self):
        pass

    def guardar_datos_excel(self, datos, nombre_archivo, con_clases):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        if con_clases:
            # Encabezados para datos con clases
            encabezados = ["Dato", "M", "F", "FA", "FR", "FAR"]
        else:
            # Encabezados para datos sin clases
            encabezados = ["Dato", "F", "FA", "FR", "FAR"]

        for col_num, encabezado in enumerate(encabezados, 1):
            sheet.cell(row=1, column=col_num, value=encabezado)

        # Escribir datos
        for row_num, dato in enumerate(datos, 2):
            sheet.cell(row=row_num, column=1, value=dato[0])  # Escribir el dato en la primera columna
            for col_num, valor in enumerate(dato[1:], 2):  # Empezar desde la segunda columna
                sheet.cell(row=row_num, column=col_num, value=valor)

        workbook.save(nombre_archivo)    